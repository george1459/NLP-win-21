# Extract DOI from AER official website

import requests
from lxml import etree
import time
import openpyxl

headers={'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
'Accept-Encoding': 'gzip, deflate, br',
'Accept-Language': 'zh-CN,zh;q=0.9',
'Cache-Control': 'max-age=0',
'Connection': 'keep-alive',
'Cookie': 'PHPSESSID=3ec5f9e0c57909aa6e25265632bd3d33; __utma=13214375.1759566504.1614699683.1614699683.1614699683.1; __utmc=13214375; __utmz=13214375.1614699683.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmt=1; __utmb=13214375.5.10.1614699683',
'Host': 'www.aeaweb.org',
'Sec-Fetch-Dest': 'document',
'Sec-Fetch-Mode': 'navigate',
'Sec-Fetch-Site': 'none',
'Sec-Fetch-User': '?1',
'Upgrade-Insecure-Requests': '1',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'}
res=requests.get("https://www.aeaweb.org/journals/aer/issues", headers=headers)

cmp=etree.HTML(res.text)
month_as=cmp.xpath("//section[@class='journal-preview-group']//a")
hrefs=[]
month_year=[]
DOIs=[]
for i in range(len(month_as)):
    href="https://www.aeaweb.org"+month_as[i].xpath("./@href")[0]
    text=month_as[i].xpath("./text()")[0]
    hrefs.append(href)
    month_year.append(text)
    res=requests.get(href, headers=headers)
    cmp=etree.HTML(res.text)
    dois=cmp.xpath("//section[@class='journal-article-group']/article/@id")
    dois.pop(0)
    DOIs.append(dois)
    print(href, text, dois)
    time.sleep(3)

data=openpyxl.Workbook()
table=data.active
keys=["month", "year", "doi"]
for i in range(len(keys)):
    table.cell(1, i+1).value=keys[i]
index=0


for i in range(len(hrefs)):
    for j in range(len(DOIs[i])):
        table.cell(index+2,1).value=month_year[i].split(" ")[0]
        table.cell(index+2,2).value=month_year[i].split(" ")[1]
        table.cell(index+2,3).value=DOIs[i][j]
        index+=1

        
data.save("./DOIs.xlsx")



